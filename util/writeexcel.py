# coding:utf-8
from openpyxl import load_workbook
import openpyxl,random
from util.log.mylog import Log
from config.globalparam import data_path

def copy_excel(excelpath1,excelpath2):
    wb2=openpyxl.Workbook()
    wb2.save(excelpath2)

    #读取数据
    wb1=openpyxl.load_workbook(excelpath1)
    wb2=openpyxl.load_workbook(excelpath2)
    sheetnames=wb1.sheetnames
    sheets1=wb1.sheetnames
    sheets2=wb2.sheetnames
    sheet1=wb1[sheets1[0]]
    sheet2=wb2[sheets2[0]]
    max_row=sheet1.max_row
    max_column=sheet1.max_column

    for m in list(range(1, max_row + 1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格

    wb2.save(excelpath2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()



class Write_excel(object):
    '''修改excel数据'''

    def __init__(self, filename,):
        self.filename=filename  #定义文件名获取方法
        self.wb = load_workbook(self.filename)  #打开已有的工作簿
        self.ws = self.wb.active  # 激活sheet,得到活动表表名

    def write(self,i,row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        #得到工作簿的所有工作表
        self.sheetnames = self.wb.sheetnames
        #得到工作簿的第i个下标的sheetname
        self.ws = self.wb[self.sheetnames[i]]
        #通过行列读
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)
        Log().info('更新数据完成！')

    def remove_order(self, num, dic4, dic2, bookname, sheetname):
        flag = False
        sum = 0
        for i in dic4.keys():
            if num == i and dic2[num][9] == '已退款':  # 判定指定属性，确定删除行
                sum += 1
                dic2.pop(num)
                self.updateExcle(dic2, data_path, sheetname[0], sum)
                flag = True
                break
        return flag


    def printing(self):
        wb=openpyxl.load_workbook(data_path)
        sheets=wb.sheetnames

        for i in range(len(sheets)):
            sheet = wb[sheets[i]]
            print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
            for r in range(1, sheet.max_row + 1):
                if r == 1:
                    print('\n' + ''.join(
                        [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
                else:
                    print(''.join([str(sheet.cell(row=r, column=c).value).ljust(20) for c in range(1, sheet.max_column + 1)]))




if __name__ == '__main__':
    # copy_excel('data.xlsx','data2.xlsx')
    # Write_excel(data_path).printing()
    wb = load_workbook(data_path)
    sheetnames=wb.sheetnames
    print(sheetnames)
    a="测试"
    b='zyxwvucbuetsrqponmlkjihgfeda'
    for m in range(2,4):
        Write_excel(data_path).write(0,m,1,a + b[m])
        # Write_excel(data_path).write(1, m, 1, b[m])
        # Write_excel(data_path).write(2, m, 1, a + b[m])
        # Write_excel(data_path).write(3, m, 1, a + b[m])
        # Write_excel(data_path).write(4, m, 1, a + b[m])